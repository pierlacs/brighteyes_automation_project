#!/usr/bin/env python3
"""
export_to_excel.py  —  BrightEyes Excel Export Utility

Reads the two key CSVs from the project and exports them as a
properly formatted .xlsx workbook with two sheets:
  - Sheet 1: Purchase History (Demo 3 — supplier analysis data)
  - Sheet 2: Weekly Business Review (Demo 5 — WBR metrics)

Satisfies bingo card task: "Read / export an Excel file"

Usage:
    python3 export_to_excel.py

Output:
    output/brighteyes_data_export.xlsx
"""

from pathlib import Path
import sys

# ── Check dependencies ────────────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.cell.cell import MergedCell          # needed for the type-check fix
    import pandas as pd
except ImportError:
    print("❌ Missing dependencies. Run: pip install openpyxl pandas")
    sys.exit(1)

ROOT = Path(__file__).parent

# ── File paths ────────────────────────────────────────────────────────────────
PURCHASE_CSV = ROOT / "demo3_supplier" / "data" / "purchase_history.csv"
WBR_CSV      = ROOT / "demo5_wbr"     / "data" / "weekly_metrics.csv"
OUTPUT_DIR   = ROOT / "output"
OUTPUT_XLSX  = OUTPUT_DIR / "brighteyes_data_export.xlsx"

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# ── Style constants ───────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="1E3A5F")  # dark navy
ALT_ROW_FILL = PatternFill("solid", fgColor="EBF5FB")  # light blue
WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
NORMAL_FONT  = Font(size=10)
CENTERED     = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN   = Alignment(horizontal="left",   vertical="center")

THIN   = Side(style="thin", color="D5D8DC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_header_row(ws, row_num: int, n_cols: int) -> None:
    """Apply dark navy header style to a given row."""
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = CENTERED
        cell.border    = BORDER


def _style_data_rows(ws, start_row: int, end_row: int, n_cols: int) -> None:
    """Apply alternating row fills and borders to data rows."""
    for row in range(start_row, end_row + 1):
        fill = ALT_ROW_FILL if row % 2 == 0 else WHITE_FILL
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill      = fill
            cell.font      = NORMAL_FONT
            cell.border    = BORDER
            cell.alignment = CENTERED


def _auto_width(ws) -> None:
    """Set column widths based on content length.

    FIX: openpyxl creates MergedCell placeholder objects for merged ranges.
    These do not have a column_letter attribute and raise AttributeError
    if accessed via col[0]. We skip them explicitly with an isinstance check.
    """
    for col in ws.columns:
        # Skip columns whose first cell is a MergedCell placeholder
        if isinstance(col[0], MergedCell):
            continue
        max_len = max(
            (len(str(c.value or "")) for c in col if not isinstance(c, MergedCell)),
            default=8,
        )
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)


def write_purchase_sheet(wb) -> None:
    """Sheet 1: Purchase history with supplier risk colour-coding."""
    print("📦 Reading purchase history CSV...")
    if not PURCHASE_CSV.exists():
        print(f"⚠️  File not found: {PURCHASE_CSV}")
        return

    df = pd.read_csv(PURCHASE_CSV)
    df["quality_rating"] = pd.to_numeric(df["quality_rating"], errors="coerce")

    # Compute risk level per row — thresholds match demo3_supplier/supplier_analysis.py
    def _risk(q: float) -> str:
        if q < 3.5:  return "HIGH"
        if q < 4.0:  return "MEDIUM"
        return "LOW"

    df["risk_level"] = df["quality_rating"].apply(
        lambda q: _risk(float(q)) if pd.notna(q) else "—"
    )

    ws = wb.active
    ws.title = "Purchase History"
    ws.freeze_panes = "A3"   # freeze title + header rows when scrolling

    # Title row — plain single cell (no merge = no MergedCell issue)
    ws["A1"] = "BrightEyes — Purchase History  |  Demo 3: Supplier Analysis"
    ws["A1"].font      = Font(bold=True, size=12, color="1E3A5F")
    ws["A1"].alignment = LEFT_ALIGN
    ws.row_dimensions[1].height = 22

    # Write headers + data from row 2 onward
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    n_cols = len(df.columns)
    _style_header_row(ws, 2, n_cols)
    _style_data_rows(ws, 3, len(df) + 2, n_cols)

    # Colour-code the risk_level column
    risk_col_idx = list(df.columns).index("risk_level") + 1
    risk_colors  = {
        "HIGH":   ("E74C3C", "FFFFFF"),
        "MEDIUM": ("FFC300", "1E3A5F"),
        "LOW":    ("2ECC71", "FFFFFF"),
    }
    for row in range(3, len(df) + 3):
        cell = ws.cell(row=row, column=risk_col_idx)
        bg, fg = risk_colors.get(str(cell.value), ("FFFFFF", "000000"))
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(bold=True, color=fg, size=10)

    _auto_width(ws)
    print(f"  ✅ Sheet 'Purchase History' — {len(df)} rows, {n_cols} columns")


def write_wbr_sheet(wb) -> None:
    """Sheet 2: Weekly metrics with alert rows highlighted."""
    print("📊 Reading weekly metrics CSV...")
    if not WBR_CSV.exists():
        print(f"⚠️  File not found: {WBR_CSV}")
        return

    df = pd.read_csv(WBR_CSV)
    for col in ["revenue_eur", "csat_score", "support_tickets_open"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # Flag alert weeks — thresholds match demo5_wbr/wbr_generator.py
    df["status"] = "OK"
    if "revenue_eur" in df.columns:
        df.loc[df["revenue_eur"] == 0, "status"] = "ALERT"
    if "csat_score" in df.columns:
        df.loc[df["csat_score"] < 4.0, "status"] = "ALERT"
    if "support_tickets_open" in df.columns:
        df.loc[df["support_tickets_open"] > 8, "status"] = "ALERT"

    ws = wb.create_sheet("Weekly Business Review")
    ws.freeze_panes = "A3"

    # Title row — plain single cell
    ws["A1"] = "BrightEyes — Weekly Business Review  |  Demo 5: WBR Generator"
    ws["A1"].font      = Font(bold=True, size=12, color="1E3A5F")
    ws["A1"].alignment = LEFT_ALIGN
    ws.row_dimensions[1].height = 22

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    n_cols = len(df.columns)
    _style_header_row(ws, 2, n_cols)
    _style_data_rows(ws, 3, len(df) + 2, n_cols)

    # Highlight alert rows red, best revenue row green
    status_col_idx = list(df.columns).index("status") + 1
    best_rev_row = None
    if "revenue_eur" in df.columns:
        best_rev_row = int(df["revenue_eur"].idxmax()) + 3  # +2 offset for title+header, +1 for 1-indexing

    for row in range(3, len(df) + 3):
        status_val = ws.cell(row=row, column=status_col_idx).value
        if status_val == "ALERT":
            for col in range(1, n_cols + 1):
                ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="FDECEA")
            ws.cell(row=row, column=status_col_idx).font = Font(bold=True, color="E74C3C", size=10)
        elif row == best_rev_row:
            for col in range(1, n_cols + 1):
                ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="EAFAF1")
            ws.cell(row=row, column=status_col_idx).value = "BEST WEEK"
            ws.cell(row=row, column=status_col_idx).font  = Font(bold=True, color="27AE60", size=10)

    _auto_width(ws)
    print(f"  ✅ Sheet 'Weekly Business Review' — {len(df)} rows, {n_cols} columns")


# ── Main ──────────────────────────────────────────────────────────────────────
def main() -> int:
    """Entry point — builds the workbook and saves it."""
    print("\n📄 BrightEyes Excel Export")
    print("─" * 40)

    wb = openpyxl.Workbook()
    write_purchase_sheet(wb)
    write_wbr_sheet(wb)

    wb.save(OUTPUT_XLSX)
    print(f"\n✅ Saved → {OUTPUT_XLSX.relative_to(ROOT)}")
    print("   Two sheets: 'Purchase History' and 'Weekly Business Review'")
    print("   Open in Excel or Numbers to verify the formatting.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
